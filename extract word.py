import os.path
import torch
from transformers import BertTokenizer, BertModel
import pandas as pd
from openpyxl import Workbook,load_workbook

class WRITE_TO_EXCEL():
    def __init__(self,path):
        self.path = path

    def write_to_file(self,value):
        df = value
        emp = pd.DataFrame()
        if not os.path.exists(self.path):
            workbook = Workbook()
            emp.to_excel(self.path)
            workbook.create_sheet('Sheet1')
        else:
            workbook = load_workbook(self.path)
        df1 = pd.DataFrame(pd.read_excel(self.path, sheet_name='Sheet1'))
        writer = pd.ExcelWriter(self.path)
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df_rows = df1.shape[0]
        df.to_excel(writer, sheet_name='Sheet1', startrow=df_rows + 1, index=False,
                    header=False)
        writer.save()  # Save


class read_vocab:
    def __init__(self):
        self.filename = r"Google/uncased_L-12_H-768_A-12/vocab.txt" # file name

    def read(self):
        with open(self.filename, 'rb') as f:
            lines = f.read().splitlines() # read a line
        i = 0 # read from the second line
        dict = {}
        list_all_english_word = []
        while i < len(lines): # determine whether is the end of file
            if lines[i].isalpha() or '[PAD]' in lines[i].decode() or '[UNK]' in lines[i].decode()\
                    or '[CLS]' in lines[i].decode() or '[SEP]' in lines[i].decode()\
                    or '[MASK]' in lines[i].decode():
                dict[i] = lines[i]
                list_all_english_word.append(i)
            i += 1
        return dict, list_all_english_word


output_file = 'BERT_vec.xlsx'
value = [['word_id', 'word', 'vector'],]
writer = WRITE_TO_EXCEL(output_file)
read_BERT = read_vocab()
model = BertModel.from_pretrained('bert-base-uncased',
                                  output_hidden_states=True,
                                  )
# Put the model in "evaluation" mode, meaning feed-forward operation.
model.eval()
BERT_dict, BERT_dict_key = read_BERT.read()

for i in range(len(BERT_dict_key)):
    text = BERT_dict[BERT_dict_key[i]]
    marked_text = "[CLS] " + text.decode() + " [SEP]"
    tokenizer = BertTokenizer.from_pretrained('bert-base-uncased')
    tokenized_text = tokenizer.tokenize(marked_text)
    index_of_word = tokenizer.convert_tokens_to_ids(tokenized_text)
    print(tokenized_text)
    id = [1]*len(tokenized_text)
    input_tensor = torch.tensor([index_of_word])
    segment_tensor = torch.tensor([id])

    with torch.no_grad():
        output = model(input_tensor, segment_tensor)
        vec = output[2][12][0][1] #2 , layer(0-12), batch, which word,
        [['word_id', 'word', 'vector'], ]
        data = {'word_id':BERT_dict_key[i],'word':tokenized_text[1],'vector':[vec.numpy().T]}
        df = pd.DataFrame(data)
        writer.write_to_file(df)
print('Done')


